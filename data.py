#!/usr/bin/python
# -*- coding: UTF-8 -*-
import cookielib
import urllib
import urllib2
from bs4 import BeautifulSoup

import re 

from openpyxl.workbook import Workbook

#values:= a dictionary containing info for request
#data = urllib.urlencode(values)
#request = urllib2.Request(url,data)
url="http://umji.sjtu.edu.cn"
request=urllib2.Request(url)
response=urllib2.urlopen(request)

soup=BeautifulSoup(response.read(),"html.parser",from_encoding="utf-8")
links=soup.find_all('a')
link_dict={}
link_dict[url]=[]

for link in links:
    if(re.match("http://",link['href'])==None):
        continue
    if(link['href'] in link_dict[url]):
         continue
    link_dict[url].append(link['href'])

link_dict[url].remove("http://facebook.com/umsjtuji")
link_dict[url].remove("http://zaymi-bistro.ru")

data_range=link_dict[url]
data_range.append(url)
print(data_range)

for web in data_range:
    web_request=urllib2.Request(web)
    web_response=urllib2.urlopen(web_request)
    soup=BeautifulSoup(web_response.read(),"html.parser",from_encoding="urf-8")
    web_links=soup.find_all('a')
    link_dict[web]=[]
    for web_link in web_links:
        if(not web_link.has_attr('href')):
             continue
        if(web_link['href'] not in data_range):
             continue
        if(re.match("http://",web_link['href'])==None):
             continue
        if(web_link['href'] in link_dict[web]):
             continue
        link_dict[web].append(web_link['href'])
    print('outbound links of url:'+web)
    print(len(link_dict[web]))

data={}
for web in data_range:
     data[web]=[]
     for outbound in data_range:
          if outbound in link_dict[web]:
               data[web].append(1)
          else:
               data[web].append(0)
     data[web].append(len(link_dict[web]))

print(data)

wb=Workbook()
ws=wb.create_sheet()
for x in range(1,41):
     for y in range(1,42):
          cell=ws.cell(row=x,column=y)
          cell.value=data[data_range[x-1]][y-1]

for y in range(1,41):
     cell=ws.cell(row=0,column=y)
     cell.value=data_range[y-1]
     cell=ws.cell(row=y,column=0)
     cell.value=data_range[y-1]
wb.save('data.xlsx')
